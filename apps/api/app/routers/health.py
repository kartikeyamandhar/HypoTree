import os
import uuid
import logging
from fastapi import APIRouter, UploadFile, File
from fastapi.responses import JSONResponse

logger = logging.getLogger(__name__)
router = APIRouter(tags=["health"])

UPLOAD_DIR = "/tmp/hypotree_uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.get("/health")
async def health():
    return {"status": "ok", "version": "0.7.0"}


@router.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Accept PDF, CSV, Excel, or image uploads."""
    file_id = str(uuid.uuid4())[:8]
    ext = os.path.splitext(file.filename or "")[1].lower()
    allowed = {".pdf", ".csv", ".xlsx", ".xls", ".png", ".jpg", ".jpeg"}

    if ext not in allowed:
        return JSONResponse(status_code=400, content={"detail": f"Unsupported file type: {ext}"})

    path = os.path.join(UPLOAD_DIR, f"{file_id}{ext}")
    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)

    file_info = {
        "id": file_id,
        "filename": file.filename,
        "extension": ext,
        "size_bytes": len(content),
        "path": path,
    }

    # Parse based on type
    parsed = _parse_file(path, ext)
    file_info["parsed"] = parsed

    logger.info("Uploaded %s (%d bytes) -> %s", file.filename, len(content), path)
    return file_info


def _parse_file(path: str, ext: str) -> dict:
    """Extract structured content from uploaded file."""
    if ext == ".pdf":
        return _parse_pdf(path)
    elif ext == ".csv":
        return _parse_csv(path)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(path)
    elif ext in (".png", ".jpg", ".jpeg"):
        return {"type": "image", "note": "Image uploaded. Vision analysis available in future phase."}
    return {}


def _parse_pdf(path: str) -> dict:
    try:
        import pymupdf
        doc = pymupdf.open(path)
        pages = []
        for i, page in enumerate(doc):
            text = page.get_text()
            pages.append({"page": i + 1, "text": text[:2000], "chars": len(text)})
        return {"type": "pdf", "pages": len(doc), "content": pages[:10]}
    except Exception as e:
        return {"type": "pdf", "error": str(e)}


def _parse_csv(path: str) -> dict:
    try:
        import csv
        with open(path, "r") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            rows = []
            for i, row in enumerate(reader):
                if i >= 5: break
                rows.append(row)
        return {"type": "csv", "headers": headers, "sample_rows": rows, "column_count": len(headers)}
    except Exception as e:
        return {"type": "csv", "error": str(e)}


def _parse_excel(path: str) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = []
        for name in wb.sheetnames[:3]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 6: break
                rows.append([str(c) if c is not None else "" for c in row])
            sheets.append({"name": name, "rows": rows})
        return {"type": "excel", "sheets": sheets}
    except Exception as e:
        return {"type": "excel", "error": str(e)}
